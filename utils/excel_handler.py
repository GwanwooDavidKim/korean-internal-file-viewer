# -*- coding: utf-8 -*-
"""
Excel 파일 처리 모듈 (Excel File Handler)

pandas와 openpyxl을 사용하여 Excel 파일을 읽고 표시합니다.
"""
import pandas as pd
import openpyxl
from typing import List, Dict, Any, Optional, Tuple
import os


class ExcelHandler:
    """
    Excel 파일 처리를 위한 클래스입니다.
    
    주요 기능:
    - Excel 시트 목록 조회
    - 시트 데이터 읽기
    - 셀 범위 읽기
    - Excel 메타데이터 조회
    """
    
    def __init__(self):
        """ExcelHandler 인스턴스를 초기화합니다."""
        self.supported_extensions = ['.xlsx', '.xlsm']  # .xls, .xlsb는 추가 엔진 필요
        self.max_rows = 100   # 표시할 최대 행 수 (성능 최적화)
        self.max_cols = 20    # 표시할 최대 열 수 (성능 최적화)
        self.preview_rows = 50  # 미리보기용 더 적은 행 수
    
    def can_handle(self, file_path: str) -> bool:
        """
        파일이 이 핸들러가 처리할 수 있는 형식인지 확인합니다.
        
        Args:
            file_path (str): 파일 경로
            
        Returns:
            bool: 처리 가능 여부
        """
        return any(file_path.lower().endswith(ext) for ext in self.supported_extensions)
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Excel 파일의 모든 시트 이름을 반환합니다.
        
        Args:
            file_path (str): Excel 파일 경로
            
        Returns:
            List[str]: 시트 이름 목록
        """
        try:
            # openpyxl을 사용하여 시트 이름 조회 (더 안정적)
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
            
        except Exception as e:
            print(f"시트 이름 조회 오류 ({file_path}): {e}")
            return []
    
    def read_sheet(self, file_path: str, sheet_name: Optional[str] = None, 
                   max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> Dict[str, Any]:
        """
        지정된 시트의 데이터를 읽습니다.
        
        Args:
            file_path (str): Excel 파일 경로
            sheet_name (Optional[str]): 시트 이름 (None이면 첫 번째 시트)
            max_rows (Optional[int]): 최대 행 수
            max_cols (Optional[int]): 최대 열 수
            
        Returns:
            Dict[str, Any]: 시트 데이터와 메타정보
        """
        try:
            if max_rows is None:
                max_rows = self.max_rows
            if max_cols is None:
                max_cols = self.max_cols
            
            # pandas로 Excel 읽기
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows)
            else:
                df = pd.read_excel(file_path, nrows=max_rows)
            
            # 열 수 제한
            if len(df.columns) > max_cols:
                df = df.iloc[:, :max_cols]
                cols_truncated = True
            else:
                cols_truncated = False
            
            # NaN 값을 빈 문자열로 대체
            df = df.fillna('')
            
            # 데이터를 딕셔너리 형태로 변환
            data = df.to_dict('records')
            
            return {
                'data': data,
                'columns': list(df.columns),
                'row_count': len(df),
                'col_count': len(df.columns),
                'total_rows': len(df),  # 실제로는 전체 행 수를 알기 어려움
                'cols_truncated': cols_truncated,
                'rows_truncated': len(df) >= max_rows,
                'sheet_name': sheet_name,
            }
            
        except Exception as e:
            return {
                'error': f"시트 읽기 오류: {e}",
                'data': [],
                'columns': [],
                'row_count': 0,
                'col_count': 0,
            }
    
    def get_preview_data(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Excel 파일의 미리보기 데이터를 반환합니다. (빠른 로딩을 위해 최적화)
        
        Args:
            file_path (str): Excel 파일 경로
            sheet_name (Optional[str]): 시트 이름 (None이면 첫 번째 시트)
            
        Returns:
            Dict[str, Any]: 미리보기 데이터
        """
        return self.read_sheet(file_path, sheet_name, max_rows=self.preview_rows, max_cols=self.max_cols)
    
    def get_sheet_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        시트의 상세 정보를 반환합니다.
        
        Args:
            file_path (str): Excel 파일 경로
            sheet_name (str): 시트 이름
            
        Returns:
            Dict[str, Any]: 시트 정보
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            if sheet_name not in workbook.sheetnames:
                return {'error': f"시트 '{sheet_name}'을 찾을 수 없습니다"}
            
            worksheet = workbook[sheet_name]
            
            # 시트 차원 정보
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 사용된 범위의 실제 데이터 확인
            actual_max_row = 0
            actual_max_col = 0
            
            for row in range(1, min(max_row + 1, 100)):  # 처음 100행만 확인
                for col in range(1, min(max_col + 1, 50)):  # 처음 50열만 확인
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        actual_max_row = max(actual_max_row, row)
                        actual_max_col = max(actual_max_col, col)
            
            workbook.close()
            
            return {
                'sheet_name': sheet_name,
                'max_row': max_row,
                'max_col': max_col,
                'actual_max_row': actual_max_row,
                'actual_max_col': actual_max_col,
                'estimated_data_rows': actual_max_row,
                'estimated_data_cols': actual_max_col,
            }
            
        except Exception as e:
            return {'error': f"시트 정보 조회 오류: {e}"}
    
    def get_cell_value(self, file_path: str, sheet_name: str, 
                      row: int, col: int) -> Optional[Any]:
        """
        특정 셀의 값을 반환합니다.
        
        Args:
            file_path (str): Excel 파일 경로
            sheet_name (str): 시트 이름
            row (int): 행 번호 (1부터 시작)
            col (int): 열 번호 (1부터 시작)
            
        Returns:
            Optional[Any]: 셀 값 또는 None
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            worksheet = workbook[sheet_name]
            
            cell_value = worksheet.cell(row=row, column=col).value
            
            workbook.close()
            return cell_value
            
        except Exception as e:
            print(f"셀 값 읽기 오류 ({file_path}, {sheet_name}, {row}, {col}): {e}")
            return None
    
    def get_workbook_info(self, file_path: str) -> Dict[str, Any]:
        """
        Excel 워크북의 전체 정보를 반환합니다.
        
        Args:
            file_path (str): Excel 파일 경로
            
        Returns:
            Dict[str, Any]: 워크북 정보
        """
        try:
            if not os.path.exists(file_path):
                return {'error': '파일을 찾을 수 없습니다'}
            
            # 파일 기본 정보
            file_size = os.path.getsize(file_path)
            
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            # 시트 정보 수집
            sheets_info = []
            for sheet_name in workbook.sheetnames:
                sheet_info = self.get_sheet_info(file_path, sheet_name)
                if 'error' not in sheet_info:
                    sheets_info.append(sheet_info)
            
            workbook.close()
            
            return {
                'filename': os.path.basename(file_path),
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2),
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'sheets_info': sheets_info,
            }
            
        except Exception as e:
            return {'error': f"워크북 정보 조회 오류: {e}"}
    
    def search_in_sheet(self, file_path: str, sheet_name: str, 
                       search_term: str, max_results: int = 20) -> List[Dict[str, Any]]:
        """
        시트에서 특정 텍스트를 검색합니다.
        
        Args:
            file_path (str): Excel 파일 경로
            sheet_name (str): 시트 이름
            search_term (str): 검색할 텍스트
            max_results (int): 최대 결과 수
            
        Returns:
            List[Dict[str, Any]]: 검색 결과 목록
        """
        try:
            results = []
            search_term = search_term.lower()
            
            # pandas로 데이터 읽기
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # 각 셀에서 검색
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    if pd.notna(value) and search_term in str(value).lower():
                        results.append({
                            'row': row_idx + 2,  # Excel은 1부터, 헤더 고려해서 +2
                            'column': col_idx + 1,
                            'column_name': df.columns[col_idx],
                            'value': str(value),
                            'context': str(value)[:100] + ('...' if len(str(value)) > 100 else ''),
                        })
                        
                        if len(results) >= max_results:
                            return results
            
            return results
            
        except Exception as e:
            print(f"시트 검색 오류 ({file_path}, {sheet_name}): {e}")
            return []